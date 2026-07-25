import os
import re
import argparse
from pathlib import Path
import pandas as pd

def parse_gaze_log(file_path):
    """
    解析單一 txt 檔案的視線追蹤紀錄。
    涵蓋 Stage 1~10，包含：
    1. 特徵擴充與重新排序：於反應等級前加入 'total分數'。
    2. 次級事件回溯 (Fallback Mechanism) 處理 TH#2 碰撞。
    """
    vid = file_path.stem
    text = file_path.read_text(encoding='utf-8')
        
    vid_match = re.search(r'Video:.*[/\\](\d+)\.mp4', text)
    if vid_match:
        vid = vid_match.group(1)

    parts = text.split('=== Full Event Log ===')
    summary_text = parts[0]
    event_log_text = parts[1] if len(parts) > 1 else ""

    # 定義 Stage Schema (新增 total分數，置於反應等級之前)
    # 🌟 新增：計分結束時間欄位（Stage 1-7=換牌瞬間；8/10=T0+10s；9=畫好了+3s or T0+15s）
    stages_schema = {
        1:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        2:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        3:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        4:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        5:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        6:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        7:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        8:  ['T0', '計分結束時間', 'TH', 'Pointing次數', '第一次Pointing', 'TH次數', 'total分數', '反應等級'],  # 🌟 Stage 8 指向特規
        9:  ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級'],
        10: ['T0', '計分結束時間', 'TB', 'TH', 'Pointing次數', '第一次Pointing', 'TB次數', 'TH次數', 'total分數', '反應等級']
    }
    
    row_data = {'影片編號': vid, '狀態': '完成'}
    
    for stage, cols in stages_schema.items():
        # 變數初始化，設定預設空值為 'x'
        level, total_score, t0, tb, th = 'x', 'x', 'x', 'x', 'x'
        pt_cnt, tb_cnt, th_cnt = 0, 0, 0
        pt_first_time = 'x'  # 🌟 新增：第一次 Pointing 時間（無則 'x'）
        score_end_time = 'x'  # 🌟 新增：計分結束時間

        # 更新後的正則表達式：同時捕獲等級 (group 1) 與分數 (group 2)
        # 範例匹配: 反應等級 LR(1分)
        block_pattern = rf"Stage\s+{stage}\s+--.*?反應等級\s+([A-Za-z]+)\s*\((\d+)分?\).*?(?=Stage\s+\d+\s+--|={40}|$)"
        match = re.search(block_pattern, summary_text, re.DOTALL)

        # 🌟 修改：文字檔中完全沒有此 Stage 的區塊標題 → 整組欄位標示「未偵測」
        # 與「有偵測到但無反應（T0 存在、TB/TH 為 x / F）」明確區分
        if not re.search(rf"Stage\s+{stage}\s+--", summary_text):
            for field in cols:
                row_data[f'Stage{stage}_{field}'] = '未偵測'
            continue

        if match:
            level = match.group(1)       # 提取英文字母，如 LR, LI, HI
            total_score = match.group(2) # 提取數字，如 1, 2, 3
            block = match.group(0)
            
            t0_m = re.search(r'T0\s*=\s*(.*?)\n', block)
            if t0_m:
                val = t0_m.group(1).strip()
                if 'not achieved' not in val.lower() and '--' not in val:
                    t0 = val.split()[0]
                    
            tb_m = re.search(r'TB\s*=\s*(.*?)\n', block)
            if tb_m:
                val = tb_m.group(1).strip()
                if 'not achieved' not in val.lower() and '--' not in val:
                    tb = val.split()[0]
                    cnt_m = re.search(r'x(\d+)', val)
                    if cnt_m: tb_cnt = int(cnt_m.group(1))
                    
            th_m = re.search(r'TH\s*=\s*(.*?)\n', block)
            if th_m:
                val = th_m.group(1).strip()
                if 'not achieved' not in val.lower() and '--' not in val:
                    th = val.split()[0]
                    cnt_m = re.search(r'x(\d+)', val)
                    if cnt_m: th_cnt = int(cnt_m.group(1))
                    
            pt_m = re.search(r'Pointing\s*=\s*(.*?)\n', block)
            if pt_m:
                val = pt_m.group(1).strip()
                if 'not detected' not in val.lower() and '--' not in val:
                    cnt_m = re.search(r'x(\d+)', val)
                    if cnt_m: pt_cnt = int(cnt_m.group(1))
                    # 🌟 新增：提取第一次 Pointing 的絕對時間（格式：12.34s (+...from T0) x3）
                    time_m = re.search(r'([\d.]+)s', val)
                    if time_m: pt_first_time = time_m.group(1) + 's'

            # 🌟 新增：計分結束時間（格式：計分結束 = 123.45s）
            end_m = re.search(r'計分結束\s*=\s*([\d.]+)s', block)
            if end_m:
                score_end_time = end_m.group(1) + 's'

        # ----------------------------------------------------
        # 次級事件回溯邏輯 (Fallback Mechanism)
        # ----------------------------------------------------
        stage_log_pattern = rf"Stage change -> {stage}\b(.*?)(?=Stage change ->|$)"
        collision = False
        
        if stage != 8:
            if tb != 'x' and th != 'x' and tb == th:
                collision = True
            elif tb == 'x' and t0 != 'x' and th != 'x' and t0 == th:
                collision = True
        else:
            if t0 != 'x' and th != 'x' and t0 == th:
                collision = True

        if collision:
            stage_log_match = re.search(stage_log_pattern, event_log_text, re.DOTALL)
            if stage_log_match:
                stage_log = stage_log_match.group(1)
                th2_m = re.search(r'\[\s*([\d\.]+)s\s*\]\s*TH#2\b', stage_log)
                if th2_m:
                    th = th2_m.group(1) + 's'

        # ----------------------------------------------------
        # 特徵映射與記憶體指派 (嚴格遵守 Schema 定義的排序)
        # ----------------------------------------------------
        if stage != 8:
            row_data[f'Stage{stage}_T0'] = t0
            row_data[f'Stage{stage}_計分結束時間'] = score_end_time  # 🌟 新增
            row_data[f'Stage{stage}_TB'] = tb
            row_data[f'Stage{stage}_TH'] = th
            row_data[f'Stage{stage}_Pointing次數'] = pt_cnt
            row_data[f'Stage{stage}_第一次Pointing'] = pt_first_time
            row_data[f'Stage{stage}_TB次數'] = tb_cnt
            row_data[f'Stage{stage}_TH次數'] = th_cnt
            row_data[f'Stage{stage}_total分數'] = total_score
            row_data[f'Stage{stage}_反應等級'] = level
        else: # Stage 8
            row_data[f'Stage{stage}_T0'] = t0
            row_data[f'Stage{stage}_計分結束時間'] = score_end_time  # 🌟 新增
            row_data[f'Stage{stage}_TH'] = th
            row_data[f'Stage{stage}_Pointing次數'] = pt_cnt
            row_data[f'Stage{stage}_第一次Pointing'] = pt_first_time
            row_data[f'Stage{stage}_TH次數'] = th_cnt
            row_data[f'Stage{stage}_total分數'] = total_score
            row_data[f'Stage{stage}_反應等級'] = level
            
    return row_data

def write_formatted_excel(df, output_path):
    """
    🌟 新增：輸出格式化 Excel。
    - 第 1 列：每個 Stage 的合併標題（基本資訊 / Stage 1 ~ Stage 10）
    - 第 2 列：欄位名（去掉 Stage{n}_ 前綴，如 T0 / TB / TH / 反應等級）
    - Stage 與 Stage 之間插入一個窄空欄作為視覺分隔
    - 凍結前兩列與「影片編號/狀態」欄，捲動時保持可見
    資料內容與 CSV 完全相同，只是排版不同。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # 欄位分組：基本資訊 + 各 Stage（依 DataFrame 現有欄位自動歸組）
    groups = [("基本資訊", [c for c in ['影片編號', '狀態'] if c in df.columns])]
    for stage in range(1, 11):
        cols = [c for c in df.columns if c.startswith(f'Stage{stage}_')]
        if cols:
            groups.append((f"Stage {stage}", cols))

    wb = Workbook()
    ws = wb.active
    ws.title = "統整"

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    title_font  = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    col_idx = 1
    for gi, (title, cols) in enumerate(groups):
        start = col_idx
        # 第 1 列：Stage 標題（跨該組所有欄位合併置中）
        ws.cell(row=1, column=start, value=title)
        if len(cols) > 1:
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1, end_column=start + len(cols) - 1)
        tc = ws.cell(row=1, column=start)
        tc.font = title_font; tc.alignment = center; tc.fill = header_fill

        # 第 2 列：欄位名 + 第 3 列起：資料
        for j, col in enumerate(cols):
            short = col.split('_', 1)[1] if col.startswith('Stage') and '_' in col else col
            hc = ws.cell(row=2, column=start + j, value=short)
            hc.font = title_font; hc.alignment = center; hc.fill = header_fill
            for r, v in enumerate(df[col].tolist(), start=3):
                ws.cell(row=r, column=start + j, value=v)
            ws.column_dimensions[get_column_letter(start + j)].width = \
                max(10, len(str(short)) * 2 + 4)
        col_idx += len(cols)

        # 組間分隔：插入一個窄空欄
        if gi < len(groups) - 1:
            ws.column_dimensions[get_column_letter(col_idx)].width = 2
            col_idx += 1

    # 凍結：前兩列（標題）+ 基本資訊欄（影片編號/狀態 + 分隔欄）
    ws.freeze_panes = "D3"
    wb.save(output_path)


def main():
    # 🌟 修改：路徑以「腳本所在位置」為基準（summarize/），
    # 預設輸入 = summarize/files/，輸出 = summarize/output/，
    # 不論從哪個工作目錄執行都不會跑錯資料夾。
    base_dir = Path(__file__).resolve().parent

    parser = argparse.ArgumentParser(description="批次處理 Gaze Tracking txt 檔案 (含分數萃取與欄位重排)")
    parser.add_argument('-d', '--dir', type=str, default=None,
                        help="指定包含 txt 檔案的資料夾路徑（預設：summarize/files/）")
    args = parser.parse_args()

    # 🌟 修改：不再互動式詢問路徑，預設直接使用 summarize/files/
    target_dir = Path(args.dir) if args.dir else (base_dir / "files")

    if not target_dir.is_dir():
        print(f"[Error] 系統無法找到該目錄: {target_dir.resolve()}")
        return

    txt_files = list(target_dir.glob("*.txt"))
    
    if not txt_files:
        print(f"[Warning] 目錄 {target_dir.resolve()} 中未發現任何 .txt 檔案。")
        return
        
    print(f"正在處理 {len(txt_files)} 個檔案...")
    
    all_rows = []
    for f_path in txt_files:
        try:
            row = parse_gaze_log(f_path)
            all_rows.append(row)
            print(f"  - 成功解析: {f_path.name}")
        except Exception as e:
            print(f"  - 解析失敗: {f_path.name}, 錯誤訊息: {e}")
            
    # 🌟 新增：全部解析失敗時提前結束，避免空 DataFrame 排序時 KeyError
    if not all_rows:
        print("[Error] 沒有任何檔案解析成功，不輸出表格。")
        return

    df = pd.DataFrame(all_rows)

    # 強制將編號轉為數值以進行穩定排序 (Stable Sort)
    def extract_numeric_key(series):
        return series.apply(
            lambda x: int(re.search(r'\d+', str(x)).group()) if re.search(r'\d+', str(x)) else float('inf')
        )
    
    df = df.sort_values(by='影片編號', key=extract_numeric_key).reset_index(drop=True)
    
    # 🌟 修改：輸出表格改存到 summarize/output/（不存在就自動建立）
    output_dir = base_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "txt_資料統整.csv"
    df.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f"\n[Success] 資料轉換與排序處理完成！結果已存至:\n{output_path.resolve()}")

    # 🌟 新增：另存格式化 Excel（Stage 標題列 + 組間分隔欄），資料與 CSV 相同
    xlsx_path = output_dir / "txt_資料統整_1.xlsx"
    try:
        write_formatted_excel(df, xlsx_path)
        print(f"[Success] 格式化 Excel（含 Stage 標題與分隔）已存至:\n{xlsx_path.resolve()}")
    except ImportError:
        print("[Warning] 未安裝 openpyxl，略過 Excel 輸出（pip install openpyxl）")
    except Exception as e:
        print(f"[Warning] Excel 輸出失敗（CSV 不受影響）：{e}")

if __name__ == "__main__":
    main()
